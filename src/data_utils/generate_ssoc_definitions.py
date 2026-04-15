from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SOURCE_PATH = PROJECT_ROOT / "data" / "ssoc2024-detailed-definitions.xlsx"
OUTPUT_PATH = PROJECT_ROOT / "data" / "ssoc2024-detailed-definitions.csv"
HEADER_ROW = 5


def main() -> None:
    # The source workbook has a small header offset, so extract only the two columns
    # the runtime pipeline actually needs.
    workbook = load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    header = []
    for value in rows[HEADER_ROW - 1][:2]:
        header.append("" if value is None else str(value).strip())

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)  # data/ already exists
    with OUTPUT_PATH.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(header)
        for row in rows[HEADER_ROW:]:
            writer.writerow([
                "" if row[0] is None else row[0],
                "" if row[1] is None else row[1],
            ])

    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
