from __future__ import annotations

import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = PROJECT_ROOT / "data"

FRAMEWORK_PATH = DATA_DIR / "jobsandskills-skillsfuture-skills-framework-dataset.xlsx"
UNIQUE_SKILLS_PATH = DATA_DIR / "jobsandskills-skillsfuture-unique-skills-list.xlsx"
UNIQUE_MAPPING_PATH = DATA_DIR / "jobsandskills-skillsfuture-tsc-to-unique-skills-mapping.xlsx"

OUTPUT_PATH = DATA_DIR / "skillsfuture_mapping.csv"

FRAMEWORK_SHEET = "TSC_CCS_Key"
UNIQUE_SKILLS_SHEET = "Unique Skills List"
UNIQUE_MAPPING_SHEET = "TSC to Unique Skill Mapping"


def _clean_text(value: object) -> str:
    text = "" if value is None else str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def _clean_token(value: object) -> str:
    text = _clean_text(value).lower()
    text = re.sub(r"[^a-z0-9\s\+\#]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _truthy(value: object) -> bool:
    return _clean_text(value).lower() == "true"


def _channel_from_type(skill_type: object) -> str:
    return "transferable" if _clean_token(skill_type) == "ccs" else "technical"


def _open_sheet(path: Path, sheet_name: str):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    return workbook, sheet


def _load_parent_sector_hints() -> dict[str, list[str]]:
    # The unique-skills workbook lacks some sector context, so recover it from the
    # parent-skill mapping sheet first.
    workbook, sheet = _open_sheet(UNIQUE_MAPPING_PATH, UNIQUE_MAPPING_SHEET)
    try:
        rows = sheet.iter_rows(values_only=True)
        header = ["" if cell is None else str(cell) for cell in next(rows)]
        index = {name: i for i, name in enumerate(header)}

        sector_map: dict[str, set[str]] = defaultdict(set)
        for row in rows:
            parent_skill = _clean_text(row[index["parent_skill_title"]])
            sector = _clean_text(row[index["sector_title"]])
            if parent_skill and sector:
                sector_map[parent_skill].add(sector)

        return {skill: sorted(sectors) for skill, sectors in sector_map.items()}
    finally:
        workbook.close()


def _build_structured_rows() -> list[dict[str, str]]:
    rows_by_skill: dict[str, dict[str, str]] = {}

    framework_book, framework_sheet = _open_sheet(FRAMEWORK_PATH, FRAMEWORK_SHEET)
    try:
        rows = framework_sheet.iter_rows(values_only=True)
        header = ["" if cell is None else str(cell) for cell in next(rows)]
        index = {name: i for i, name in enumerate(header)}

        for row in rows:
            title = _clean_text(row[index["TSC_CCS Title"]])
            skill_norm = _clean_token(title)
            if not skill_norm:
                continue

            sector = _clean_text(row[index["Sector"]])
            cluster = _clean_text(row[index["TSC_CCS Category"]]) or sector
            description = _clean_text(row[index["TSC_CCS Description"]])
            code = _clean_text(row[index["TSC Code"]])
            skill_type = row[index["TSC_CCS Type"]]

            note_parts = [part for part in [sector, cluster, description, code] if part]
            candidate = {
                "skill_norm": skill_norm,
                "channel": _channel_from_type(skill_type),
                "framework_cluster": cluster,
                "skillsfuture_note": " | ".join(note_parts),
            }

            existing = rows_by_skill.get(skill_norm)
            if existing is None:
                rows_by_skill[skill_norm] = candidate
                continue

            if existing["channel"] != "transferable" and candidate["channel"] == "transferable":
                existing["channel"] = candidate["channel"]
            if not existing["framework_cluster"] and candidate["framework_cluster"]:
                existing["framework_cluster"] = candidate["framework_cluster"]
            if candidate["skillsfuture_note"] and len(candidate["skillsfuture_note"]) > len(existing["skillsfuture_note"]):
                existing["skillsfuture_note"] = candidate["skillsfuture_note"]
    finally:
        framework_book.close()

    # Start from the richer framework extract, then backfill remaining unique skills
    # that do not already appear there.
    parent_sector_hints = _load_parent_sector_hints()

    unique_book, unique_sheet = _open_sheet(UNIQUE_SKILLS_PATH, UNIQUE_SKILLS_SHEET)
    try:
        rows = unique_sheet.iter_rows(values_only=True)
        header = ["" if cell is None else str(cell) for cell in next(rows)]
        index = {name: i for i, name in enumerate(header)}

        for row in rows:
            parent_skill = _clean_text(row[index["parent_skill_title"]])
            skill_norm = _clean_token(parent_skill)
            if not skill_norm or skill_norm in rows_by_skill:
                continue

            description = _clean_text(row[index["parent_skill_description"]])
            skill_type = row[index["skill_type"]]
            tags = []
            if _truthy(row[index["Emerging Skills"]]):
                tags.append("Emerging Skills")
            if _truthy(row[index["CASL Skills"]]):
                tags.append("CASL Skills")

            sectors = parent_sector_hints.get(parent_skill, [])
            cluster = " / ".join(tags) if tags else ("Multi-sector" if sectors else "")

            note_parts = []
            if sectors:
                note_parts.append("Sectors: " + ", ".join(sectors[:5]))
            if description:
                note_parts.append(description)
            if tags:
                note_parts.append("Tags: " + ", ".join(tags))

            rows_by_skill[skill_norm] = {
                "skill_norm": skill_norm,
                "channel": _channel_from_type(skill_type),
                "framework_cluster": cluster,
                "skillsfuture_note": " | ".join(note_parts),
            }
    finally:
        unique_book.close()

    return [rows_by_skill[key] for key in sorted(rows_by_skill)]


def _write_structured_extract() -> int:
    fieldnames = ["skill_norm", "channel", "framework_cluster", "skillsfuture_note"]
    rows = _build_structured_rows()

    with OUTPUT_PATH.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    return len(rows)


def main() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    row_count = _write_structured_extract()
    print(f"Wrote {OUTPUT_PATH} with {row_count} rows")


if __name__ == "__main__":
    main()
